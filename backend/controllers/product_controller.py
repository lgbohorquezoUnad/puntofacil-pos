import os
from datetime import datetime
from io import BytesIO

from flask import current_app, jsonify, request, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from werkzeug.utils import secure_filename

from models.product_model import (
    create_product,
    delete_product,
    get_categories,
    get_or_create_category,
    get_product_by_barcode,
    get_product_by_id,
    get_products,
    update_product,
    update_product_image,
)
from models.sale_model import get_daily_summary

ALLOWED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
TEMPLATE_HEADERS = [
    "producto_id",
    "nombre",
    "codigo_barras",
    "categoria",
    "precio",
    "stock",
    "stock_minimo",
    "imagen_url",
]



def get_current_user_context():
    identity = get_jwt_identity()
    claims = get_jwt() or {}

    if isinstance(identity, dict):
        return {
            "id": identity.get("id"),
            "nombre": identity.get("nombre") or claims.get("nombre"),
            "rol": identity.get("rol") or claims.get("rol"),
        }

    user_id = identity
    try:
        user_id = int(identity)
    except (TypeError, ValueError):
        pass

    return {
        "id": user_id,
        "nombre": claims.get("nombre"),
        "rol": claims.get("rol"),
    }



def require_admin_user():
    current_user = get_current_user_context()
    if current_user.get("rol") != "admin":
        return None, (jsonify({"error": "No autorizado"}), 403)
    return current_user, None



def serialize_products(products):
    result = []
    for p in products:
        result.append(
            {
                "id": p[0],
                "nombre": p[1],
                "precio": float(p[2]),
                "stock": p[3],
                "stock_minimo": p[4],
                "categoria": p[5],
                "categoria_id": p[6],
                "codigo_barras": p[7],
                "imagen_url": p[8] if len(p) > 8 else None,
            }
        )
    return result



def parse_product_payload(data):
    try:
        category_id = int(data.get("categoria_id"))
        stock = int(data.get("stock"))
        stock_minimo = int(data.get("stock_minimo", 5))
        price = float(data.get("precio"))
    except (TypeError, ValueError):
        raise ValueError("Categoria, stock, stock minimo o precio invalidos")

    if stock < 0 or stock_minimo < 0 or price < 0:
        raise ValueError("Stock, stock minimo y precio no pueden ser negativos")

    nombre = (data.get("nombre") or "").strip()
    codigo_barras = (data.get("codigo_barras") or "").strip()
    imagen_url = (data.get("imagen_url") or "").strip() or None

    if not nombre:
        raise ValueError("El nombre del producto es obligatorio")

    return {
        "nombre": nombre,
        "codigo_barras": codigo_barras,
        "precio": price,
        "stock": stock,
        "stock_minimo": stock_minimo,
        "categoria_id": category_id,
        "imagen_url": imagen_url,
    }



def normalize_excel_header(value):
    return str(value or "").strip().lower().replace(" ", "_")



def build_inventory_template(categories):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Productos"
    sheet.append(TEMPLATE_HEADERS)
    sheet.append(["", "Cafe 500g", "7701234567890", "Bebidas", 18500, 20, 5, "https://ejemplo.com/cafe.jpg"])

    header_fill = PatternFill(fill_type="solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)

    for index, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = header_fill
        cell.font = header_font
        sheet.column_dimensions[cell.column_letter].width = 20

    notes = workbook.create_sheet("Instrucciones")
    notes.append(["Campo", "Detalle"])
    notes.append(["producto_id", "Opcional. Si lo envias, actualiza ese producto existente."])
    notes.append(["nombre", "Obligatorio."])
    notes.append(["codigo_barras", "Opcional. Si coincide con un producto, se actualiza."])
    notes.append(["categoria", "Obligatoria. Si no existe, el sistema la crea."])
    notes.append(["precio", "Obligatorio. Numero decimal."])
    notes.append(["stock", "Obligatorio. Numero entero."])
    notes.append(["stock_minimo", "Opcional. Por defecto 5."])
    notes.append(["imagen_url", "Opcional. URL publica de la imagen."])
    notes.append(["Categorias actuales", ", ".join(category["nombre"] for category in categories) or "Sin categorias registradas"])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream



def save_uploaded_product_image(product_id, uploaded_file):
    original_name = secure_filename(uploaded_file.filename or "")
    extension = os.path.splitext(original_name)[1].lower()

    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise ValueError("Formato de imagen no permitido. Usa PNG, JPG, JPEG, WEBP o GIF")

    uploads_dir = os.path.join(current_app.static_folder, "uploads", "products")
    os.makedirs(uploads_dir, exist_ok=True)

    filename = f"product_{product_id}_{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}{extension}"
    absolute_path = os.path.join(uploads_dir, filename)
    uploaded_file.save(absolute_path)
    return f"/static/uploads/products/{filename}"



def parse_excel_number(raw_value, field_name, row_number, integer=False, default=None):
    if raw_value in (None, ""):
        if default is not None:
            return default
        raise ValueError(f"Fila {row_number}: el campo {field_name} es obligatorio")

    try:
        value = int(raw_value) if integer else float(raw_value)
    except (TypeError, ValueError):
        raise ValueError(f"Fila {row_number}: el campo {field_name} no es valido")

    if value < 0:
        raise ValueError(f"Fila {row_number}: el campo {field_name} no puede ser negativo")

    return value



def import_workbook_rows(mysql, workbook):
    sheet = workbook.active
    headers = [normalize_excel_header(cell.value) for cell in sheet[1]]
    missing_headers = [header for header in ["nombre", "categoria", "precio", "stock"] if header not in headers]
    if missing_headers:
        raise ValueError("La plantilla no contiene las columnas obligatorias: " + ", ".join(missing_headers))

    header_map = {header: index for index, header in enumerate(headers)}
    created = 0
    updated = 0
    errors = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = {header: row[index] if index < len(row) else None for header, index in header_map.items()}
        if all(value in (None, "") for value in values.values()):
            continue

        try:
            product_id = values.get("producto_id")
            product_id = int(product_id) if product_id not in (None, "") else None
            nombre = str(values.get("nombre") or "").strip()
            categoria_nombre = str(values.get("categoria") or "").strip()
            codigo_barras = str(values.get("codigo_barras") or "").strip()
            imagen_url = str(values.get("imagen_url") or "").strip() or None
            precio = parse_excel_number(values.get("precio"), "precio", row_number)
            stock = parse_excel_number(values.get("stock"), "stock", row_number, integer=True)
            stock_minimo = parse_excel_number(values.get("stock_minimo"), "stock_minimo", row_number, integer=True, default=5)

            if not nombre:
                raise ValueError(f"Fila {row_number}: el nombre es obligatorio")
            if not categoria_nombre:
                raise ValueError(f"Fila {row_number}: la categoria es obligatoria")

            category = get_or_create_category(mysql, categoria_nombre)

            existing = None
            if product_id:
                existing = get_product_by_id(mysql, product_id)
                if existing is None:
                    raise ValueError(f"Fila {row_number}: el producto_id {product_id} no existe")
            elif codigo_barras:
                existing = get_product_by_barcode(mysql, codigo_barras)

            if existing is not None:
                current = serialize_products([existing])[0]
                update_product(
                    mysql,
                    existing[0],
                    nombre=nombre,
                    codigo_barras=codigo_barras,
                    precio=precio,
                    stock=stock,
                    categoria_id=category["id"],
                    stock_minimo=stock_minimo,
                    imagen_url=imagen_url if imagen_url is not None else current.get("imagen_url"),
                )
                updated += 1
            else:
                create_product(
                    mysql,
                    nombre=nombre,
                    codigo_barras=codigo_barras,
                    precio=precio,
                    stock=stock,
                    categoria_id=category["id"],
                    stock_minimo=stock_minimo,
                    imagen_url=imagen_url,
                )
                created += 1
        except ValueError as error:
            errors.append(str(error))

    return {
        "created": created,
        "updated": updated,
        "errors": errors,
        "processed": created + updated,
    }


@jwt_required()
def list_products(mysql):
    products = get_products(mysql)
    return jsonify(serialize_products(products))


@jwt_required()
def list_categories(mysql):
    return jsonify(get_categories(mysql))


@jwt_required()
def create_product_controller(mysql):
    current_user, admin_error = require_admin_user()
    if admin_error:
        return admin_error

    data = request.json or {}

    try:
        payload = parse_product_payload(data)
        product_id = create_product(mysql, **payload)
        product = get_product_by_id(mysql, product_id)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    except Exception:
        return jsonify({"error": "No fue posible crear el producto"}), 500

    return jsonify({"message": "Producto creado", "product": serialize_products([product])[0]}), 201


@jwt_required()
def update_product_controller(mysql, product_id):
    current_user, admin_error = require_admin_user()
    if admin_error:
        return admin_error

    data = request.json or {}

    try:
        payload = parse_product_payload(data)
        updated = update_product(mysql, product_id, **payload)
        if not updated:
            return jsonify({"error": "Producto no encontrado"}), 404
        product = get_product_by_id(mysql, product_id)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    except Exception:
        return jsonify({"error": "No fue posible actualizar el producto"}), 500

    return jsonify({"message": "Producto actualizado", "product": serialize_products([product])[0]}), 200


@jwt_required()
def update_product_image_controller(mysql, product_id):
    current_user, admin_error = require_admin_user()
    if admin_error:
        return admin_error

    product = get_product_by_id(mysql, product_id)
    if product is None:
        return jsonify({"error": "Producto no encontrado"}), 404

    try:
        if "image" in request.files:
            uploaded_file = request.files["image"]
            if not uploaded_file or not uploaded_file.filename:
                return jsonify({"error": "Debes adjuntar una imagen valida"}), 400
            image_url = save_uploaded_product_image(product_id, uploaded_file)
        else:
            data = request.json or {}
            image_url = (data.get("image_url") or "").strip() or None

        updated = update_product_image(mysql, product_id, image_url)
        if not updated:
            return jsonify({"error": "Producto no encontrado"}), 404
        updated_product = get_product_by_id(mysql, product_id)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    except Exception:
        return jsonify({"error": "No fue posible actualizar la imagen del producto"}), 500

    return jsonify({"message": "Imagen actualizada", "product": serialize_products([updated_product])[0]}), 200


@jwt_required()
def download_product_import_template_controller(mysql):
    current_user, admin_error = require_admin_user()
    if admin_error:
        return admin_error

    stream = build_inventory_template(get_categories(mysql))
    return send_file(
        stream,
        as_attachment=True,
        download_name="plantilla_carga_productos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@jwt_required()
def import_products_controller(mysql):
    current_user, admin_error = require_admin_user()
    if admin_error:
        return admin_error

    uploaded_file = request.files.get("file")
    if uploaded_file is None or not uploaded_file.filename:
        return jsonify({"error": "Debes adjuntar un archivo Excel"}), 400

    if not uploaded_file.filename.lower().endswith(".xlsx"):
        return jsonify({"error": "El archivo debe estar en formato .xlsx"}), 400

    try:
        workbook = load_workbook(uploaded_file, data_only=True)
        summary = import_workbook_rows(mysql, workbook)
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    except Exception:
        return jsonify({"error": "No fue posible procesar el archivo de carga masiva"}), 500

    return jsonify({
        "message": "Carga masiva procesada",
        "summary": summary,
    }), 200


@jwt_required()
def delete_product_controller(mysql, product_id):
    current_user, admin_error = require_admin_user()
    if admin_error:
        return admin_error

    try:
        deleted = delete_product(mysql, product_id)
        if not deleted:
            return jsonify({"error": "Producto no encontrado"}), 404
    except ValueError as error:
        return jsonify({"error": str(error)}), 400
    except Exception:
        return jsonify({"error": "No fue posible eliminar el producto"}), 500

    return jsonify({"message": "Producto eliminado"}), 200


@jwt_required()
def get_admin_dashboard(mysql):
    current_user, admin_error = require_admin_user()
    if admin_error:
        return admin_error

    return jsonify(get_daily_summary(mysql)), 200
